import io
import os
import re
from urllib.parse import quote_plus

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

# =========================================================
# HEATMAP WMS V16 — VISÃO 2D + 3D + MAPA CLÁSSICO EVOLUÍDO
# ---------------------------------------------------------
# Baseado na V13 do Mapa 3D e nos conceitos operacionais
# do Swap Inteligente V21.
# =========================================================

st.set_page_config(
    page_title="Heatmap WMS V16",
    page_icon="🔥",
    layout="wide",
)

# -----------------------------
# SEGREDOS / AMBIENTE
# -----------------------------
def _get_secret_or_env(key: str, default=None):
    try:
        if key in st.secrets:
            value = st.secrets[key]
            if value not in [None, ""]:
                return str(value)
    except Exception:
        pass

    try:
        mysql_cfg = st.secrets.get("mysql", {})
        mysql_map = {
            "WMS_DB_HOST": "host",
            "WMS_DB_PORT": "port",
            "WMS_DB_NAME": "database",
            "WMS_DB_USER": "user",
            "WMS_DB_PASS": "password",
        }
        mysql_key = mysql_map.get(key)
        if mysql_key and mysql_key in mysql_cfg:
            value = mysql_cfg[mysql_key]
            if value not in [None, ""]:
                return str(value)
    except Exception:
        pass

    value = os.getenv(key, default)
    if value in [None, ""]:
        return default
    return str(value)


@st.cache_resource
def get_engine():
    host = _get_secret_or_env("WMS_DB_HOST", "127.0.0.1")
    port = _get_secret_or_env("WMS_DB_PORT", "3306")
    db = _get_secret_or_env("WMS_DB_NAME", "wms_apanha")
    user = _get_secret_or_env("WMS_DB_USER", "wms_user")
    password_raw = _get_secret_or_env("WMS_DB_PASS", "")

    if password_raw == "":
        raise RuntimeError(
            "WMS_DB_PASS vazio no ambiente/streamlit secrets. "
            "Configure a senha do banco no secrets.toml ou nas variáveis de ambiente."
        )

    password = quote_plus(password_raw)
    url = f"mysql+pymysql://{user}:{password}@{host}:{port}/{db}"
    return create_engine(url, pool_pre_ping=True)


def test_connection():
    with get_engine().connect() as conn:
        conn.execute(text("SELECT 1"))


# -----------------------------
# CARGA DE DADOS
# -----------------------------
@st.cache_data(ttl=300, show_spinner=False)
def load_empresas():
    sql = text(
        """
        SELECT DISTINCT nroempresa
        FROM vw_mapa_3d_cd
        WHERE nroempresa IS NOT NULL
        ORDER BY nroempresa
        """
    )
    with get_engine().connect() as conn:
        df = pd.read_sql(sql, conn)

    if df.empty:
        return []
    return df["nroempresa"].dropna().astype("int64").tolist()


@st.cache_data(ttl=300, show_spinner=False)
def load_data(nroempresa: int):
    sql = text(
        """
        SELECT
            nroempresa,
            dep,
            rua,
            predio,
            nivel,
            apto,
            sala,
            seqendereco,
            codproduto,
            descproduto,
            qtdatual,
            capacidade,
            perc_ocupacao,
            COALESCE(visitas, 0) AS visitas,
            COALESCE(qtd_saida, 0) AS qtd_saida,
            COALESCE(media_dia_cx, 0) AS media_dia_cx,
            COALESCE(dias_sem_reposicao, 0) AS dias_sem_reposicao,
            COALESCE(curva_endereco, '') AS curva_endereco,
            COALESCE(curva_reposicao, '') AS curva_reposicao,
            COALESCE(mapa_de_calor, 'SEM INFO') AS mapa_de_calor,
            COALESCE(gap_curva, 0) AS gap_curva,
            COALESCE(curva_correta, 'SEM INFO') AS curva_correta,
            COALESCE(justificativa, '') AS justificativa,
            COALESCE(acao_recomendada, '') AS acao_recomendada,
            COALESCE(flag_urgente_lt1d, 0) AS flag_urgente_lt1d,
            COALESCE(status_endereco, 'LIVRE') AS status_endereco
        FROM vw_mapa_3d_cd
        WHERE nroempresa = :nroempresa
        ORDER BY rua, predio, nivel, apto, codproduto
        """
    )
    with get_engine().connect() as conn:
        return pd.read_sql(sql, conn, params={"nroempresa": nroempresa})


# -----------------------------
# AUXILIARES
# -----------------------------
def norm_text(x):
    if x is None:
        return ""
    return str(x).strip()


def zfill_numeric_text(x, width: int):
    s = norm_text(x)
    if s == "":
        return ""
    return s.zfill(width) if s.isdigit() else s


def natural_sort_key(value):
    s = norm_text(value)
    if s == "":
        return (1, "")
    parts = re.split(r"(\d+)", s)
    out = []
    for p in parts:
        if p == "":
            continue
        out.append(int(p) if p.isdigit() else p.upper())
    # Retorna tupla hashable para uso seguro em pandas.sort_values(key=...)
    return (0, tuple(out))


def ordenar_natural(valores, reverse: bool = False):
    return sorted([v for v in valores if norm_text(v) != ""], key=natural_sort_key, reverse=reverse)


def fmt_posicao(valor):
    s = norm_text(valor)
    if s == "":
        return ""
    try:
        f = float(s.replace(",", "."))
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return s


def posicao_chave(predio, sala):
    return f"{fmt_posicao(predio)}|{fmt_posicao(sala)}"


def label_posicao(predio, sala):
    p = fmt_posicao(predio)
    s = fmt_posicao(sala)
    return f"{p}.{s}" if s != "" else p


def normalizar_pqr(valor):
    s = norm_text(valor).upper()
    if s == "":
        return "SEM PQR"
    for letra in ["P", "Q", "R"]:
        if s.startswith(letra):
            return letra
    return "SEM PQR"


def cor_pqr(letra: str) -> str:
    return {
        "P": "#22C55E",
        "Q": "#FACC15",
        "R": "#EF4444",
        "SEM PQR": "#94A3B8",
    }.get(normalizar_pqr(letra), "#94A3B8")


def safe_numeric_series(s: pd.Series, default=0):
    return pd.to_numeric(s, errors="coerce").fillna(default)


def classificar_ocupacao(perc):
    try:
        perc = float(perc)
    except Exception:
        return "SEM INFO"

    if perc <= 0:
        return "VAZIO"
    if perc <= 40:
        return "BAIXA"
    if perc <= 80:
        return "MEDIA"
    return "ALTA"


def cor_ocupacao(cat):
    return {
        "VAZIO": "#BDBDBD",
        "BAIXA": "#64B5F6",
        "MEDIA": "#FFD54F",
        "ALTA": "#E53935",
        "SEM INFO": "#9E9E9E",
    }.get(cat, "#9E9E9E")


def cor_mapa_calor(cat):
    base = norm_text(cat).upper()
    if base.startswith("P"):
        return "#22C55E"
    if base.startswith("Q"):
        return "#FACC15"
    if base.startswith("R"):
        return "#EF4444"
    return "#94A3B8"


def cor_curva(cat):
    return {
        "P": "#22C55E",
        "Q": "#FACC15",
        "R": "#EF4444",
        "SEM INFO": "#94A3B8",
        "": "#94A3B8",
    }.get(norm_text(cat).upper(), "#94A3B8")


def normalizar_status(cat):
    s = norm_text(cat).upper()
    if s in ["LIVRE", "L", "FREE", "VAZIO", "VAGO"]:
        return "LIVRE"
    if s in ["OCUPADO", "OCUP", "O", "CHEIO"]:
        return "OCUPADO"
    if s in ["RESERVADO", "RESERV", "RES", "RSV", "R"]:
        return "RESERVADO"
    if s in ["BLOQUEADO", "BLOQ", "BLQ"]:
        return "BLOQUEADO"
    if s in ["AVARIADO", "AVARIA", "AVR"]:
        return "AVARIADO"
    if s in ["DIVERGENTE", "DIVERG", "DIV"]:
        return "DIVERGENTE"
    if s in ["", "NONE", "NAN", "NULL", "SEM INFO", "SEM_STATUS", "SEM STATUS"]:
        return "SEM STATUS"
    return "SEM STATUS"


def cor_status(cat):
    return {
        "LIVRE": "#22C55E",
        "OCUPADO": "#EF4444",
        "RESERVADO": "#2563EB",
        "BLOQUEADO": "#8E24AA",
        "AVARIADO": "#8B4513",
        "DIVERGENTE": "#FB8C00",
        "SEM STATUS": "#94A3B8",
    }.get(normalizar_status(cat), "#94A3B8")


def normalizar_tipo_divergencia(valor):
    s = norm_text(valor).upper()
    if s == "":
        return "SEM BASE"
    validos = {"OK", "DIVERGENTE", "PLAN06_MAIS_RECENTE", "OPERACAO_MAIS_RECENTE", "SEM_MOV_OPERACIONAL", "SEM_PLAN06", "SEM BASE"}
    return s if s in validos else "SEM BASE"


def cor_confiabilidade(tipo: str) -> str:
    tipo = normalizar_tipo_divergencia(tipo)
    return {
        "OK": "#22C55E",
        "PLAN06_MAIS_RECENTE": "#F59E0B",
        "OPERACAO_MAIS_RECENTE": "#2563EB",
        "DIVERGENTE": "#EF4444",
        "SEM_MOV_OPERACIONAL": "#7C3AED",
        "SEM_PLAN06": "#0EA5E9",
        "SEM BASE": "#94A3B8",
    }.get(tipo, "#94A3B8")


STATUSS_ATIVOS_MAPA = ["LIVRE", "OCUPADO", "RESERVADO", "BLOQUEADO", "AVARIADO", "DIVERGENTE"]


def endereco_ativo_mask(df: pd.DataFrame) -> pd.Series:
    if df.empty:
        return pd.Series([], dtype=bool)

    status_ok = df.get("STATUS_ENDERECO", pd.Series(index=df.index, dtype=object)).fillna("").astype(str).str.upper().isin(STATUSS_ATIVOS_MAPA)
    seq_ok = pd.to_numeric(df.get("seqendereco", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0).gt(0)
    qtd_ok = pd.to_numeric(df.get("qtdatual", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0).gt(0)
    cod_ok = pd.to_numeric(df.get("codproduto", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0).gt(0)
    perc_ok = pd.to_numeric(df.get("perc_ocupacao", pd.Series(index=df.index, dtype=float)), errors="coerce").fillna(0).gt(0)

    rua_ok = df.get("RUA_TXT", pd.Series(index=df.index, dtype=object)).fillna("").astype(str).str.strip().ne("")
    predio_ok = df.get("PREDIO_TXT", pd.Series(index=df.index, dtype=object)).fillna("").astype(str).str.strip().ne("")
    nivel_ok = df.get("NIVEL_TXT", pd.Series(index=df.index, dtype=object)).fillna("").astype(str).str.strip().ne("")
    sala_ok = df.get("SALA_TXT", pd.Series(index=df.index, dtype=object)).fillna("").astype(str).str.strip().ne("")

    estrutura_ok = rua_ok & predio_ok & nivel_ok & sala_ok
    return estrutura_ok & (status_ok | seq_ok | qtd_ok | cod_ok | perc_ok)


def normalizar_dados(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for c in [
        "nroempresa", "rua", "predio", "nivel", "apto", "sala", "seqendereco",
        "codproduto", "qtdatual", "capacidade", "perc_ocupacao", "visitas", "qtd_saida",
        "media_dia_cx", "dias_sem_reposicao", "gap_curva", "flag_urgente_lt1d"
    ]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    for c in ["descproduto", "justificativa", "acao_recomendada"]:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].fillna("").astype(str)

    for origem, destino in [
        ("curva_endereco", "CURVA_ENDERECO"),
        ("curva_reposicao", "CURVA_REPOSICAO"),
        ("mapa_de_calor", "MAPA_DE_CALOR"),
        ("curva_correta", "CURVA_CORRETA"),
        ("status_endereco", "STATUS_ENDERECO"),
    ]:
        if origem not in df.columns:
            df[origem] = ""
        df[destino] = (
            df[origem]
            .fillna("")
            .astype(str)
            .str.upper()
            .replace({"NAN": "", "NONE": ""})
        )

    df["STATUS_ENDERECO"] = df["STATUS_ENDERECO"].apply(normalizar_status)
    df["MAPA_DE_CALOR"] = df["MAPA_DE_CALOR"].replace({"": "SEM INFO"})
    df["CURVA_CORRETA"] = df["CURVA_CORRETA"].replace({"": "SEM INFO"})
    df["FAIXA_OCUPACAO"] = df["perc_ocupacao"].fillna(0).apply(classificar_ocupacao)
    df["FLAG_URGENTE_LT1D"] = df["flag_urgente_lt1d"].fillna(0).astype(int).eq(1)

    df["DEP_TXT"] = df["dep"].fillna("").astype(str).str.strip().apply(lambda x: zfill_numeric_text(x, 2))
    df["RUA_TXT"] = df["rua"].apply(lambda x: zfill_numeric_text(x, 3) if pd.notna(x) else "")
    df["PREDIO_TXT"] = df["predio"].apply(lambda x: zfill_numeric_text(x, 2) if pd.notna(x) else "")
    df["NIVEL_TXT"] = df["nivel"].apply(lambda x: zfill_numeric_text(x, 2) if pd.notna(x) else "")
    df["SALA_TXT"] = df["sala"].apply(lambda x: zfill_numeric_text(x, 2) if pd.notna(x) else "")

    df["X"] = df["rua"].fillna(0)
    df["Y"] = df["predio"].fillna(0)
    df["Z"] = df["nivel"].fillna(0)

    df["ENDERECO_LABEL"] = (
        df["DEP_TXT"].fillna("") + "-" +
        df["RUA_TXT"].fillna("") + "-" +
        df["PREDIO_TXT"].fillna("") + "-" +
        df["NIVEL_TXT"].fillna("") + "-" +
        df["SALA_TXT"].fillna("")
    )


    if "TIPO_DIVERGENCIA" not in df.columns:
        df["TIPO_DIVERGENCIA"] = "SEM BASE"
    df["TIPO_DIVERGENCIA"] = df["TIPO_DIVERGENCIA"].apply(normalizar_tipo_divergencia)
    for c in ["DELTA_DIAS_SAIDA", "DIAS_PLAN06", "DIAS_MOV_OPERACIONAL"]:
        if c not in df.columns:
            df[c] = np.nan

    return df


def endereco_invalido_mask(df: pd.DataFrame) -> pd.Series:
    dep_blank = df["DEP_TXT"].fillna("").eq("")
    rua_blank = df["RUA_TXT"].fillna("").eq("")
    predio_blank = df["PREDIO_TXT"].fillna("").eq("")
    nivel_blank = df["NIVEL_TXT"].fillna("").eq("")
    sala_blank = df["SALA_TXT"].fillna("").eq("")

    tipo_rec = (
        df["STATUS_ENDERECO"].fillna("").astype(str).str.upper().str.contains("REC", na=False)
        | df["ENDERECO_LABEL"].str.contains("REC", na=False)
        | df["descproduto"].fillna("").str.upper().str.contains("REC", na=False)
    )
    return dep_blank | rua_blank | predio_blank | nivel_blank | sala_blank | tipo_rec


def escolher_cor(row, modo):
    if modo == "Mapa de Giro":
        return cor_mapa_calor(row["MAPA_DE_CALOR"])
    if modo == "Curva Reposição":
        return cor_curva(row["CURVA_REPOSICAO"])
    if modo == "Curva Endereço":
        return cor_curva(row["CURVA_ENDERECO"])
    if modo == "Curva Correta":
        if row["CURVA_CORRETA"] == "SIM":
            return "#22C55E"
        if row["CURVA_CORRETA"] == "NAO":
            return "#EF4444"
        return "#94A3B8"
    if modo == "Status":
        return cor_status(row["STATUS_ENDERECO"])
    if modo == "Ocupação":
        return cor_ocupacao(row["FAIXA_OCUPACAO"])
    if modo == "Pressão Operacional":
        score = float(row.get("PRESSAO_OPERACIONAL", 0) or 0)
        if score >= 200:
            return "#B71C1C"
        if score >= 80:
            return "#E65100"
        if score >= 20:
            return "#F9A825"
        return "#2E7D32"
    if modo == "Confiabilidade de Dados":
        return cor_confiabilidade(row.get("TIPO_DIVERGENCIA", "SEM BASE"))
    return "#90A4AE"


def montar_hover(df: pd.DataFrame) -> pd.Series:
    gap_str = df["gap_curva"].fillna(0).round(0).astype("Int64").astype(str)
    return (
        "<b>Endereço:</b> " + df["ENDERECO_LABEL"].fillna("") + "<br>"
        + "<b>Produto:</b> " + df["codproduto"].fillna(0).round(0).astype("Int64").astype(str) + " - " + df["descproduto"].fillna("") + "<br>"
        + "<b>DEP:</b> " + df["DEP_TXT"] + " | <b>Rua:</b> " + df["RUA_TXT"] + " | <b>Prédio:</b> " + df["PREDIO_TXT"] + " | <b>Nível:</b> " + df["NIVEL_TXT"] + " | <b>Sala:</b> " + df["SALA_TXT"] + "<br>"
        + "<b>Visitas:</b> " + df["visitas"].fillna(0).round(0).astype("Int64").astype(str) + "<br>"
        + "<b>Qtd Saída:</b> " + df["qtd_saida"].fillna(0).round(2).astype(str) + "<br>"
        + "<b>Média Dia CX:</b> " + df["media_dia_cx"].fillna(0).round(2).astype(str) + "<br>"
        + "<b>Dias sem reposição:</b> " + df["dias_sem_reposicao"].fillna(0).round(2).astype(str) + "<br>"
        + "<b>Curva Reposição:</b> " + df["CURVA_REPOSICAO"].fillna("") + "<br>"
        + "<b>Curva Endereço:</b> " + df["CURVA_ENDERECO"].fillna("") + "<br>"
        + "<b>Gap Curva:</b> " + gap_str + "<br>"
        + "<b>Curva Correta:</b> " + df["CURVA_CORRETA"].fillna("") + "<br>"
        + "<b>Pressão:</b> " + df["PRESSAO_OPERACIONAL"].fillna(0).round(1).astype(str) + "<br>"
        + "<b>Justificativa:</b> " + df["justificativa"].fillna("") + "<br>"
        + "<b>Ação:</b> " + df["acao_recomendada"].fillna("") + "<br>"
        + "<b>Urgente &lt;1 dia:</b> " + df["FLAG_URGENTE_LT1D"].apply(lambda x: "SIM" if x else "NÃO") + "<br>"
        + "<b>Status:</b> " + df["STATUS_ENDERECO"].fillna("") + "<br>"
        + "<b>Fonte PLAN06:</b> " + df.get("ULTIMA_SAIDA_PLAN06_TXT", pd.Series([""] * len(df), index=df.index)).fillna("").astype(str) + "<br>"
        + "<b>Últ. mov. operacional:</b> " + df.get("ULTIMA_MOV_OPERACIONAL_TXT", pd.Series([""] * len(df), index=df.index)).fillna("").astype(str) + "<br>"
        + "<b>Tipo divergência:</b> " + df.get("TIPO_DIVERGENCIA", pd.Series(["SEM BASE"] * len(df), index=df.index)).fillna("").astype(str) + "<br>"
        + "<b>Delta dias:</b> " + pd.to_numeric(df.get("DELTA_DIAS_SAIDA", pd.Series([np.nan] * len(df), index=df.index)), errors="coerce").fillna(-1).astype(int).replace({-1: 0}).astype(str)
    )


def aplicar_pareto(df: pd.DataFrame, percentual: float) -> pd.DataFrame:
    if df.empty or percentual >= 1:
        out = df.copy()
        out["PARETO_ACUM"] = np.nan
        return out

    out = df.copy()
    total = out["visitas"].fillna(0).sum()
    if total <= 0:
        out["PARETO_ACUM"] = np.nan
        return out

    out = out.sort_values(["visitas", "qtd_saida"], ascending=[False, False]).copy()
    out["PARETO_ACUM"] = out["visitas"].fillna(0).cumsum() / total
    return out[out["PARETO_ACUM"] <= percentual].copy()


def obter_camera(vista_3d: str):
    if vista_3d == "Frontal":
        return dict(eye=dict(x=0.0, y=-2.6, z=0.8))
    if vista_3d == "Superior":
        return dict(eye=dict(x=0.0, y=0.0, z=2.8))
    if vista_3d == "Lateral":
        return dict(eye=dict(x=2.6, y=0.0, z=0.8))
    if vista_3d == "Isométrica":
        return dict(eye=dict(x=1.6, y=-1.4, z=1.2))
    if vista_3d == "Corredor":
        return dict(eye=dict(x=0.25, y=-3.2, z=0.45))
    return dict(eye=dict(x=1.5, y=-1.5, z=1.1))


def aplicar_layout_3d(df: pd.DataFrame, layout_mode: str) -> pd.DataFrame:
    out = df.copy()
    out["X_RENDER"] = out["X"].astype(float)
    out["Y_RENDER"] = out["Y"].astype(float)
    out["Z_RENDER"] = out["Z"].astype(float)

    if layout_mode == "Explodido por Nível":
        out["Z_RENDER"] = out["Z_RENDER"] * 1.35
    elif layout_mode == "Explodido por Prédio":
        out["Y_RENDER"] = out["Y_RENDER"] * 1.18
    elif layout_mode == "Explodido Completo":
        out["Y_RENDER"] = out["Y_RENDER"] * 1.20
        out["Z_RENDER"] = out["Z_RENDER"] * 1.40
    return out


def montar_figura_3d(df: pd.DataFrame, modo: str, vista_3d: str, layout_mode: str, marker_size: int, opacity: float):
    df = aplicar_layout_3d(df, layout_mode)

    if df.empty:
        fig = go.Figure()
        fig.update_layout(
            height=760,
            title="Mapa 3D do CD",
            annotations=[dict(text="Nenhum dado para exibir", x=0.5, y=0.5, xref="paper", yref="paper", showarrow=False, font=dict(size=18))],
        )
        return fig

    df = df.copy()
    df["COR"] = df.apply(lambda row: escolher_cor(row, modo), axis=1)
    hover_text = montar_hover(df)

    normal = df[~df["FLAG_URGENTE_LT1D"]].copy()
    urgente = df[df["FLAG_URGENTE_LT1D"]].copy()
    incorretos = df[df["CURVA_CORRETA"] == "NAO"].copy()

    fig = go.Figure()

    if not normal.empty:
        fig.add_trace(
            go.Scatter3d(
                x=normal["X_RENDER"],
                y=normal["Y_RENDER"],
                z=normal["Z_RENDER"],
                mode="markers",
                marker=dict(
                    size=marker_size,
                    color=normal["COR"],
                    opacity=opacity,
                    symbol="square",
                    line=dict(width=1, color="#0F172A"),
                ),
                text=hover_text.loc[normal.index],
                name="Mapa",
                hovertemplate="%{text}<extra></extra>",
            )
        )

    if not incorretos.empty:
        fig.add_trace(
            go.Scatter3d(
                x=incorretos["X_RENDER"],
                y=incorretos["Y_RENDER"],
                z=incorretos["Z_RENDER"],
                mode="markers",
                marker=dict(
                    size=max(marker_size + 3, 10),
                    color=incorretos["COR"],
                    opacity=1.0,
                    symbol="circle-open",
                    line=dict(width=4, color="#FB8C00"),
                ),
                text=hover_text.loc[incorretos.index],
                name="Curva incorreta",
                hovertemplate="%{text}<extra></extra>",
            )
        )

    if not urgente.empty:
        fig.add_trace(
            go.Scatter3d(
                x=urgente["X_RENDER"],
                y=urgente["Y_RENDER"],
                z=urgente["Z_RENDER"],
                mode="markers",
                marker=dict(
                    size=max(marker_size + 5, 12),
                    color=urgente["COR"],
                    opacity=1.0,
                    symbol="diamond",
                    line=dict(width=2, color="#111827"),
                ),
                text=hover_text.loc[urgente.index],
                name="Urgente <1 dia",
                hovertemplate="%{text}<extra></extra>",
            )
        )

    fig.update_layout(
        height=760,
        margin=dict(l=10, r=10, t=50, b=10),
        title=f"Mapa 3D do CD — Modo: {modo} | Vista: {vista_3d} | Layout: {layout_mode}",
        scene=dict(
            xaxis_title="Rua",
            yaxis_title="Prédio",
            zaxis_title="Nível",
            camera=obter_camera(vista_3d),
            xaxis=dict(backgroundcolor="rgb(245,245,245)", gridcolor="lightgray"),
            yaxis=dict(backgroundcolor="rgb(245,245,245)", gridcolor="lightgray"),
            zaxis=dict(backgroundcolor="rgb(245,245,245)", gridcolor="lightgray"),
            aspectmode="data",
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
    )
    return fig


def metricas_resumo(df: pd.DataFrame):
    if df.empty:
        return {
            "enderecos": 0,
            "produtos": 0,
            "visitas": 0,
            "urgentes": 0,
            "incorretos": 0,
            "ocupacao_media": 0.0,
        }

    return {
        "enderecos": int(df["ENDERECO_LABEL"].nunique()),
        "produtos": int(df["codproduto"].nunique()),
        "visitas": float(df["visitas"].fillna(0).sum()),
        "urgentes": int(df["FLAG_URGENTE_LT1D"].sum()),
        "incorretos": int((df["CURVA_CORRETA"] == "NAO").sum()),
        "ocupacao_media": float(df["perc_ocupacao"].fillna(0).mean()),
    }




def diagnostico_status(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["STATUS_ENDERECO", "QTD"])
    out = df.copy()
    out["STATUS_ENDERECO"] = out["STATUS_ENDERECO"].apply(normalizar_status)
    return (
        out.groupby("STATUS_ENDERECO", as_index=False)
        .agg(QTD=("ENDERECO_LABEL", "count"))
        .sort_values(["QTD", "STATUS_ENDERECO"], ascending=[False, True])
        .reset_index(drop=True)
    )

def gerar_heatmap_rua_nivel(df: pd.DataFrame, metrica: str):
    if df.empty:
        return go.Figure()

    agg_map = {
        "visitas": ("visitas", "sum"),
        "qtd_saida": ("qtd_saida", "sum"),
        "PRESSAO_OPERACIONAL": ("PRESSAO_OPERACIONAL", "sum"),
        "perc_ocupacao": ("perc_ocupacao", "mean"),
        "contagem": ("codproduto", "count"),
    }
    col, func = agg_map[metrica]

    pt = (
        df.groupby(["RUA_TXT", "NIVEL_TXT"], as_index=False)
        .agg(valor=(col, func))
        .sort_values(["RUA_TXT", "NIVEL_TXT"])
    )

    mat = pt.pivot(index="NIVEL_TXT", columns="RUA_TXT", values="valor").fillna(0)
    mat = mat.sort_index(ascending=False)

    title_map = {
        "visitas": "Heatmap 2D — Rua x Nível | Visitas",
        "qtd_saida": "Heatmap 2D — Rua x Nível | Saída",
        "PRESSAO_OPERACIONAL": "Heatmap 2D — Rua x Nível | Pressão Operacional",
        "perc_ocupacao": "Heatmap 2D — Rua x Nível | Ocupação Média",
        "contagem": "Heatmap 2D — Rua x Nível | Itens",
    }

    fig = px.imshow(
        mat,
        text_auto=True,
        aspect="auto",
        color_continuous_scale="YlOrRd",
        labels=dict(x="Rua", y="Nível", color="Valor"),
        title=title_map[metrica],
    )
    fig.update_layout(height=540, margin=dict(l=10, r=10, t=55, b=10))
    return fig


def gerar_mapa_2d_rua_predio(df: pd.DataFrame, modo: str):
    if df.empty:
        return go.Figure()

    if modo == "Pressão Operacional":
        agg = df.groupby(["RUA_TXT", "PREDIO_TXT"], as_index=False).agg(valor=("PRESSAO_OPERACIONAL", "sum"))
        titulo = "Mapa 2D — Rua x Prédio | Pressão Operacional"
    elif modo == "Ocupação":
        agg = df.groupby(["RUA_TXT", "PREDIO_TXT"], as_index=False).agg(valor=("perc_ocupacao", "mean"))
        titulo = "Mapa 2D — Rua x Prédio | Ocupação Média"
    else:
        agg = df.groupby(["RUA_TXT", "PREDIO_TXT"], as_index=False).agg(valor=("visitas", "sum"))
        titulo = f"Mapa 2D — Rua x Prédio | {modo}"

    fig = px.scatter(
        agg,
        x="RUA_TXT",
        y="PREDIO_TXT",
        size="valor",
        color="valor",
        color_continuous_scale="Turbo",
        title=titulo,
        labels={"RUA_TXT": "Rua", "PREDIO_TXT": "Prédio", "valor": "Valor"},
    )
    fig.update_traces(marker=dict(line=dict(width=1, color="#0F172A"), sizemode="area", sizeref=max(agg["valor"].max() / 1200, 0.01)))
    fig.update_layout(height=540, margin=dict(l=10, r=10, t=55, b=10))
    return fig


def resumo_por_rua(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    rua = (
        df.groupby(["DEP_TXT", "RUA_TXT"], as_index=False)
        .agg(
            total_enderecos=("ENDERECO_LABEL", "nunique"),
            total_produtos=("codproduto", "nunique"),
            total_visitas=("visitas", "sum"),
            total_saida=("qtd_saida", "sum"),
            pressao=("PRESSAO_OPERACIONAL", "sum"),
            urgentes=("FLAG_URGENTE_LT1D", "sum"),
            incorretos=("CURVA_CORRETA", lambda s: int((s == "NAO").sum())),
            ocupacao_media=("perc_ocupacao", "mean"),
        )
        .sort_values(["pressao", "total_visitas", "urgentes"], ascending=[False, False, False])
        .reset_index(drop=True)
    )

    rua["MAPA_DE_CALOR_RUA"] = np.select(
        [
            rua["pressao"] >= rua["pressao"].quantile(0.70) if len(rua) > 1 else False,
            rua["pressao"] >= rua["pressao"].quantile(0.30) if len(rua) > 1 else False,
        ],
        ["🔴 ALTO FLUXO", "🟡 MÉDIO FLUXO"],
        default="🟢 BAIXO FLUXO",
    )
    return rua


def top_enderecos_criticos(df: pd.DataFrame, top_n=20) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    out = df.copy()
    out["PRIORIDADE"] = np.select(
        [
            (out["CURVA_REPOSICAO"] == "P") & (out["CURVA_ENDERECO"] == "R"),
            (out["CURVA_REPOSICAO"] == "P") & (out["CURVA_ENDERECO"] == "Q"),
            (out["CURVA_REPOSICAO"] == "Q") & (out["CURVA_ENDERECO"] == "R"),
        ],
        ["🔴 CRÍTICA", "🟠 ALTA", "🟡 MÉDIA"],
        default="⚪ BAIXA",
    )

    cols = [
        "ENDERECO_LABEL", "DEP_TXT", "RUA_TXT", "PREDIO_TXT", "NIVEL_TXT", "SALA_TXT",
        "codproduto", "descproduto", "visitas", "qtd_saida", "PRESSAO_OPERACIONAL",
        "CURVA_REPOSICAO", "CURVA_ENDERECO", "CURVA_CORRETA", "PRIORIDADE", "acao_recomendada"
    ]
    return out[cols].sort_values(["PRESSAO_OPERACIONAL", "visitas", "qtd_saida"], ascending=[False, False, False]).head(top_n)


def build_filter_diagnostics(df_base: pd.DataFrame, filtros: dict) -> pd.DataFrame:
    etapas = []
    atual = df_base.copy()
    etapas.append({"etapa": "Base inicial", "qtd_linhas": int(len(atual))})

    if filtros["dep"]:
        atual = atual[atual["DEP_TXT"].isin(filtros["dep"])]
        etapas.append({"etapa": "Após DEP", "qtd_linhas": int(len(atual))})

    if filtros["rua"]:
        atual = atual[atual["RUA_TXT"].isin(filtros["rua"])]
        etapas.append({"etapa": "Após Rua", "qtd_linhas": int(len(atual))})

    if filtros["predio"]:
        atual = atual[atual["PREDIO_TXT"].isin(filtros["predio"])]
        etapas.append({"etapa": "Após Prédio", "qtd_linhas": int(len(atual))})

    if filtros["nivel"]:
        atual = atual[atual["NIVEL_TXT"].isin(filtros["nivel"])]
        etapas.append({"etapa": "Após Nível", "qtd_linhas": int(len(atual))})

    if filtros["sala"]:
        atual = atual[atual["SALA_TXT"].isin(filtros["sala"])]
        etapas.append({"etapa": "Após Sala", "qtd_linhas": int(len(atual))})

    if filtros["curva_rep"]:
        atual = atual[atual["CURVA_REPOSICAO"].isin(filtros["curva_rep"])]
        etapas.append({"etapa": "Após Curva Reposição", "qtd_linhas": int(len(atual))})

    if filtros["curva_end"]:
        atual = atual[atual["CURVA_ENDERECO"].isin(filtros["curva_end"])]
        etapas.append({"etapa": "Após Curva Endereço", "qtd_linhas": int(len(atual))})

    if filtros["status"]:
        atual = atual[atual["STATUS_ENDERECO"].isin(filtros["status"])]
        etapas.append({"etapa": "Após Status", "qtd_linhas": int(len(atual))})

    if filtros.get("divergencia"):
        atual = atual[atual["TIPO_DIVERGENCIA"].isin(filtros["divergencia"])]
        etapas.append({"etapa": "Após Confiabilidade", "qtd_linhas": int(len(atual))})

    if filtros["somente_urgentes"]:
        atual = atual[atual["FLAG_URGENTE_LT1D"]]
        etapas.append({"etapa": "Após Somente Urgentes", "qtd_linhas": int(len(atual))})

    if filtros["ocultar_invalidos"]:
        atual = atual[~endereco_invalido_mask(atual)]
        etapas.append({"etapa": "Após Ocultar Inválidos/REC", "qtd_linhas": int(len(atual))})

    if filtros.get("somente_ativos", False):
        atual = atual[endereco_ativo_mask(atual)]
        etapas.append({"etapa": "Após Somente Ativos", "qtd_linhas": int(len(atual))})

    atual = aplicar_pareto(atual, filtros["pareto"])
    etapas.append({"etapa": f"Após Pareto {int(filtros['pareto']*100)}%", "qtd_linhas": int(len(atual))})

    diag = pd.DataFrame(etapas)
    diag["zerou_aqui"] = diag["qtd_linhas"].eq(0)
    return diag




def status_ocupacao_classico(row) -> str:
    status = normalizar_status(row.get("STATUS_ENDERECO", ""))
    qtdatual = pd.to_numeric(pd.Series([row.get("qtdatual", 0)]), errors="coerce").fillna(0).iloc[0]
    perc = pd.to_numeric(pd.Series([row.get("perc_ocupacao", 0)]), errors="coerce").fillna(0).iloc[0]
    codproduto = row.get("codproduto", None)

    if status in ["RESERVADO", "BLOQUEADO", "AVARIADO", "DIVERGENTE"]:
        return status
    if status == "LIVRE" and qtdatual <= 0 and perc <= 0 and not (pd.notna(codproduto) and float(codproduto) > 0):
        return "LIVRE"
    if qtdatual > 0 or perc > 0 or (pd.notna(codproduto) and float(codproduto) > 0):
        return "OCUPADO"
    if status == "SEM STATUS":
        return "SEM STATUS"
    return status


def cor_mapa_ocupacao_classico(status: str) -> str:
    return {
        "LIVRE": "#22C55E",
        "OCUPADO": "#EF4444",
        "RESERVADO": "#2563EB",
        "BLOQUEADO": "#8B5CF6",
        "AVARIADO": "#92400E",
        "DIVERGENTE": "#F59E0B",
        "SEM STATUS": "#94A3B8",
        "SEM CADASTRO": "#CBD5E1",
    }.get(normalizar_status(status) if status != "SEM CADASTRO" else status, "#94A3B8")


def preparar_mapa_ocupacao(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT", "POSICAO_CHAVE", "POSICAO_LABEL", "STATUS_MAPA", "ENDERECO_LABEL", "descproduto", "qtdatual", "perc_ocupacao", "PQR_MAPA", "PQR_FONTE"])

    base = df.copy()
    base = base[endereco_ativo_mask(base)].copy()
    if base.empty:
        return pd.DataFrame(columns=["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT", "POSICAO_CHAVE", "POSICAO_LABEL", "STATUS_MAPA", "ENDERECO_LABEL", "descproduto", "qtdatual", "perc_ocupacao", "PQR_MAPA", "PQR_FONTE"])

    base["STATUS_MAPA"] = base.apply(status_ocupacao_classico, axis=1)
    base["PQR_FONTE"] = base["CURVA_REPOSICAO"].where(base["CURVA_REPOSICAO"].fillna("").astype(str).str.strip() != "", base["CURVA_ENDERECO"])
    base["PQR_MAPA"] = base["PQR_FONTE"].apply(normalizar_pqr)
    base["PREDIO_TXT"] = base["PREDIO_TXT"].apply(fmt_posicao)
    base["SALA_TXT"] = base["SALA_TXT"].apply(fmt_posicao)
    base["POSICAO_CHAVE"] = base.apply(lambda r: posicao_chave(r.get("PREDIO_TXT",""), r.get("SALA_TXT","")), axis=1)
    base["POSICAO_LABEL"] = base.apply(lambda r: label_posicao(r.get("PREDIO_TXT",""), r.get("SALA_TXT","")), axis=1)

    prioridade = {
        "BLOQUEADO": 5,
        "RESERVADO": 4,
        "AVARIADO": 4,
        "DIVERGENTE": 4,
        "OCUPADO": 3,
        "LIVRE": 2,
        "SEM STATUS": 1,
        "SEM CADASTRO": 0,
    }
    base["PRIORIDADE_MAPA"] = base["STATUS_MAPA"].map(prioridade).fillna(0)

    base = base.sort_values(
        by=["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT", "PRIORIDADE_MAPA"],
        key=lambda col: col.map(natural_sort_key) if col.name in ["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT"] else col,
        ascending=[True, False, True, True, False],
    )

    agg = (
        base.groupby(["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT"], as_index=False)
        .agg(
            POSICAO_CHAVE=("POSICAO_CHAVE", "first"),
            POSICAO_LABEL=("POSICAO_LABEL", "first"),
            STATUS_MAPA=("STATUS_MAPA", "first"),
            ENDERECO_LABEL=("ENDERECO_LABEL", "first"),
            descproduto=("descproduto", lambda s: next((str(x) for x in s if str(x).strip()), "")),
            qtdatual=("qtdatual", "sum"),
            perc_ocupacao=("perc_ocupacao", "mean"),
            codproduto=("codproduto", lambda s: next((int(x) for x in s if pd.notna(x) and float(x) > 0), None)),
            PQR_MAPA=("PQR_MAPA", lambda s: next((normalizar_pqr(x) for x in s if normalizar_pqr(x) != "SEM PQR"), "SEM PQR")),
            PQR_FONTE=("PQR_FONTE", lambda s: next((str(x) for x in s if str(x).strip()), "")),
        )
    )
    return agg


def _coluna_zona(base: pd.DataFrame) -> str:
    for c in ["SALA_TXT", "sala", "CURVA_ENDERECO", "CURVA_REPOSICAO"]:
        if c in base.columns:
            try:
                vals = base[c].fillna("").astype(str).str.strip()
                if (vals != "").any():
                    return c
            except Exception:
                pass
    return "__ZONA_PADRAO__"


def render_mapa_ocupacao_html(df: pd.DataFrame, somente_status: str = "Todos", localizacao_busca: str = "", endereco_selecionado: str = "", modo_fluxo: str = "Nenhum", quebra_painel: str = "Geral", visao_cor: str = "Status", pqr_filtro=None):

    mapa = preparar_mapa_ocupacao(df)
    if mapa.empty:
        return "<div style='padding:16px;font-family:Arial'>Nenhum dado para exibir.</div>", pd.DataFrame()

    zona_col = _coluna_zona(df)
    base_aux = df.copy()
    if zona_col == "__ZONA_PADRAO__":
        base_aux["__ZONA_PADRAO__"] = "GERAL"
    mapa = mapa.merge(
        base_aux.groupby(["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT"], as_index=False)[zona_col].first(),
        on=["RUA_TXT", "NIVEL_TXT", "PREDIO_TXT", "SALA_TXT"], how="left"
    )
    mapa["ZONA_MAPA"] = mapa[zona_col].fillna("GERAL").astype(str).replace({"": "GERAL"})
    mapa["DEP_MAPA"] = mapa["ENDERECO_LABEL"].astype(str).str[:2]

    if somente_status == "Somente livres":
        mapa = mapa[mapa["STATUS_MAPA"].eq("LIVRE")]
    elif somente_status == "Somente ocupados":
        mapa = mapa[mapa["STATUS_MAPA"].eq("OCUPADO")]
    elif somente_status == "Somente bloqueados/avariados":
        mapa = mapa[mapa["STATUS_MAPA"].isin(["BLOQUEADO", "RESERVADO", "AVARIADO", "DIVERGENTE"])]
    elif somente_status == "Somente vazios":
        mapa = mapa[mapa["STATUS_MAPA"].isin(["LIVRE", "SEM STATUS", "SEM CADASTRO"])]

    pqr_filtro_norm = [normalizar_pqr(x) for x in (pqr_filtro or []) if normalizar_pqr(x)]
    if pqr_filtro_norm:
        mapa = mapa[mapa["PQR_MAPA"].isin(pqr_filtro_norm)]

    visao_cor_efetiva = "PQR" if pqr_filtro_norm else visao_cor

    busca = norm_text(localizacao_busca).upper()
    endereco_sel = norm_text(endereco_selecionado).upper()
    destaque_df = pd.DataFrame()
    if busca or endereco_sel:
        termo = endereco_sel if endereco_sel else busca
        destaque_df = mapa[
            mapa["ENDERECO_LABEL"].fillna("").str.upper().str.contains(termo, na=False)
            | mapa["RUA_TXT"].fillna("").str.upper().eq(termo)
            | mapa["PREDIO_TXT"].fillna("").str.upper().eq(termo)
        ].copy()

    ruas = ordenar_natural(mapa["RUA_TXT"].dropna().unique().tolist())

    if modo_fluxo == "Entrada":
        candidatos_fluxo = set(mapa.loc[mapa["STATUS_MAPA"].isin(["LIVRE", "SEM CADASTRO"]), "ENDERECO_LABEL"].astype(str).head(80).tolist())
    elif modo_fluxo == "Saída":
        top_saida = base_aux.groupby("ENDERECO_LABEL", as_index=False).agg(qtd_saida=("qtd_saida", "sum")).sort_values("qtd_saida", ascending=False).head(80)
        candidatos_fluxo = set(top_saida["ENDERECO_LABEL"].astype(str).tolist())
    else:
        candidatos_fluxo = set()

    if quebra_painel == "Por DEP":
        grupos = [(f"DEP {g}", sub.copy()) for g, sub in mapa.groupby("DEP_MAPA", sort=True)]
    elif quebra_painel == "Por Zona":
        grupos = [(f"ZONA {g}", sub.copy()) for g, sub in mapa.groupby("ZONA_MAPA", sort=True)]
    elif quebra_painel == "DEP + Zona":
        grupos = [(f"DEP {d} | ZONA {z}", sub.copy()) for (d, z), sub in mapa.groupby(["DEP_MAPA", "ZONA_MAPA"], sort=True)]
    else:
        grupos = [("VISÃO GERAL", mapa.copy())]

    css = """
    <style>
    .mapa-wrap{background:#10263f;padding:18px 18px 10px;border-radius:14px;font-family:Arial,sans-serif;color:#fff}
    .mapa-titulo{font-size:28px;font-weight:800;text-align:center;letter-spacing:.8px;margin:0 0 12px 0}
    .mapa-sub{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:12px;flex-wrap:wrap}
    .mapa-legenda{display:flex;gap:8px;flex-wrap:wrap;font-size:12px}
    .pill{display:inline-flex;align-items:center;gap:6px;background:#163555;border:1px solid rgba(255,255,255,.15);padding:6px 10px;border-radius:999px}
    .dot{width:12px;height:12px;border-radius:3px;display:inline-block}
    .rua-card{background:#163555;border-radius:12px;padding:10px 12px;margin-bottom:14px;border:1px solid rgba(255,255,255,.12)}
    .rua-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;gap:12px;flex-wrap:wrap}
    .rua-title{font-size:18px;font-weight:800;color:#FBBF24}
    .rua-stat{font-size:13px;color:#E5E7EB}
    .grid-wrap{overflow-x:auto;padding-bottom:4px}
    .nivel-row{display:flex;align-items:center;gap:6px;margin-bottom:6px;white-space:nowrap}
    .nivel-lbl{width:46px;min-width:46px;font-weight:700;color:#D1D5DB;text-align:right;padding-right:6px}
    .cell{width:34px;height:30px;min-width:34px;border-radius:4px;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff;border:1px solid rgba(255,255,255,.18)}
    .cell span{opacity:.95}
    .cell.highlight{outline:3px solid #FDE68A;box-shadow:0 0 0 2px #92400E inset}
    .resumo-geral{margin-top:12px;background:#0B1C2D;padding:10px 12px;border-radius:10px;font-size:14px;font-weight:700;text-align:center}
    .grupo-head{font-size:20px;font-weight:800;color:#93C5FD;margin:6px 0 14px;border-left:5px solid #2563EB;padding-left:10px}
    .fluxo-entrada{outline:2px solid #10B981; box-shadow:0 0 0 2px rgba(16,185,129,.35) inset}
    .fluxo-saida{outline:2px solid #F97316; box-shadow:0 0 0 2px rgba(249,115,22,.35) inset}
    </style>
    """

    parts = [css, "<div class='mapa-wrap'>", "<div class='mapa-titulo'>MAPA DE OCUPAÇÃO</div>"]
    parts.append("<div class='mapa-sub'>")
    loc = endereco_sel if endereco_sel else (busca if busca else "TODAS")
    parts.append(f"<div><strong>Localização:</strong> {loc}</div>")
    if modo_fluxo == "Entrada":
        parts.append("<div class='pill' style='background:#065F46;border-color:#10B981'><strong>ENTRADA</strong></div>")
    elif modo_fluxo == "Saída":
        parts.append("<div class='pill' style='background:#7C2D12;border-color:#F97316'><strong>SAÍDA</strong></div>")
    parts.append("<div class='mapa-legenda'>")
    if visao_cor_efetiva == "PQR":
        for nome in ["P", "Q", "R", "SEM PQR"]:
            parts.append(f"<span class='pill'><span class='dot' style='background:{cor_pqr(nome)}'></span>{nome}</span>")
    else:
        for nome in ["LIVRE", "OCUPADO", "RESERVADO", "BLOQUEADO", "AVARIADO", "DIVERGENTE", "SEM STATUS"]:
            parts.append(f"<span class='pill'><span class='dot' style='background:{cor_mapa_ocupacao_classico(nome)}'></span>{nome}</span>")
    parts.append("</div></div>")

    total_ocup = 0
    total_livre = 0
    total_valid = 0

    for grupo_titulo, grupo_df in grupos:
        if grupo_df.empty:
            continue
        parts.append(f"<div class='grupo-head'>{grupo_titulo}</div>")
        ruas_grupo = ordenar_natural(grupo_df["RUA_TXT"].dropna().unique().tolist())
        for rua in ruas_grupo:
            rua_df = grupo_df[grupo_df["RUA_TXT"].eq(rua)].copy()
            if rua_df.empty:
                continue
            ocup = int(rua_df["STATUS_MAPA"].eq("OCUPADO").sum())
            livre = int(rua_df["STATUS_MAPA"].eq("LIVRE").sum())
            valid = int(rua_df["STATUS_MAPA"].isin(["OCUPADO", "LIVRE", "RESERVADO", "BLOQUEADO", "AVARIADO", "DIVERGENTE", "SEM STATUS"]).sum())
            perc = (ocup / valid * 100.0) if valid else 0.0
            total_ocup += ocup
            total_livre += livre
            total_valid += valid

            parts.append("<div class='rua-card'>")
            parts.append(f"<div class='rua-head'><div class='rua-title'>RUA {rua}</div><div class='rua-stat'>C: {ocup} | V: {livre} | {perc:.1f}%</div></div>")
            parts.append("<div class='grid-wrap'>")
            niveis_rua = ordenar_natural(rua_df["NIVEL_TXT"].dropna().unique().tolist(), reverse=True)
            for nivel in niveis_rua:
                parts.append(f"<div class='nivel-row'><div class='nivel-lbl'>N{int(nivel) if str(nivel).isdigit() else nivel}</div>")
                nivel_df = rua_df[rua_df["NIVEL_TXT"].eq(nivel)].copy()
                nivel_df = nivel_df.sort_values(by=["PREDIO_TXT", "SALA_TXT"], key=lambda col: col.map(natural_sort_key))
                for _, row in nivel_df.iterrows():
                    predio = row.get("PREDIO_TXT", "")
                    sala = row.get("SALA_TXT", "")
                    status = row["STATUS_MAPA"]
                    pqr = normalizar_pqr(row.get("PQR_MAPA", "SEM PQR"))
                    color = cor_pqr(pqr) if visao_cor_efetiva == "PQR" else cor_mapa_ocupacao_classico(status)
                    cod = "" if pd.isna(row.get("codproduto")) else str(int(row.get("codproduto")))
                    desc = str(row.get("descproduto", ""))[:60]
                    qtd = 0 if pd.isna(row.get("qtdatual")) else float(row.get("qtdatual"))
                    ocupc = 0 if pd.isna(row.get("perc_ocupacao")) else float(row.get("perc_ocupacao"))
                    title = f"{row.get('ENDERECO_LABEL','')} | Status: {status} | PQR: {pqr} | Fonte PQR: {row.get('PQR_FONTE','')} | Prédio: {predio} | Sala: {sala} | Cod: {cod} | Produto: {desc} | Qtde: {qtd:.2f} | Ocupação: {ocupc:.1f}% | Zona: {row.get('ZONA_MAPA','GERAL')}"
                    highlight = ((busca and busca in str(row.get("ENDERECO_LABEL", "")).upper()) or (endereco_sel and endereco_sel == str(row.get("ENDERECO_LABEL", "")).upper()))
                    cls = "cell highlight" if highlight else "cell"
                    if str(row.get("ENDERECO_LABEL", "")) in candidatos_fluxo:
                        cls += " fluxo-entrada" if modo_fluxo == "Entrada" else " fluxo-saida" if modo_fluxo == "Saída" else ""
                    label = row.get("POSICAO_LABEL", label_posicao(predio, sala))
                    parts.append(f"<div class='{cls}' title='{title}' style='background:{color}'><span>{label}</span></div>")
                parts.append("</div>")
            parts.append("</div></div>")

    perc_total = (total_ocup / total_valid * 100.0) if total_valid else 0.0
    parts.append(f"<div class='resumo-geral'>Ocupação Geral: {perc_total:.1f}% ({total_ocup}/{total_valid})</div>")
    parts.append("</div>")
    return "".join(parts), destaque_df


def gerar_planilha_mapa_ocupacao(df: pd.DataFrame) -> bytes:
    mapa = preparar_mapa_ocupacao(df)
    output = io.BytesIO()
    if mapa.empty:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame({"mensagem": ["Sem dados para mapa de ocupação"]}).to_excel(writer, sheet_name="Mapa_Ocupacao", index=False)
        output.seek(0)
        return output.getvalue()

    matriz = mapa.pivot_table(index=["RUA_TXT", "NIVEL_TXT"], columns="PREDIO_TXT", values="STATUS_MAPA", aggfunc="first", fill_value="SEM CADASTRO")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        matriz.to_excel(writer, sheet_name="Mapa_Ocupacao")
        ws = writer.sheets["Mapa_Ocupacao"]
        color_map = {
            "LIVRE": "22C55E",
            "OCUPADO": "EF4444",
            "RESERVADO": "2563EB",
            "BLOQUEADO": "8B5CF6",
            "AVARIADO": "92400E",
            "DIVERGENTE": "F59E0B",
            "SEM STATUS": "94A3B8",
            "SEM CADASTRO": "CBD5E1",
        }
        for row in ws.iter_rows(min_row=2, min_col=3):
            for cell in row:
                val = str(cell.value)
                fill = PatternFill(fill_type="solid", fgColor=color_map.get(val, "FFFFFF"))
                cell.fill = fill
                if val not in ["SEM CADASTRO", "LIVRE", "SEM STATUS"]:
                    cell.font = Font(color="FFFFFF", bold=True)
        for c in ws[1]:
            c.font = Font(color="FFFFFF", bold=True)
            c.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 18)
    output.seek(0)
    return output.getvalue()



def _estilizar_cabecalho(ws):
    fill_head = PatternFill(fill_type="solid", fgColor="1F4E78")
    font_head = Font(color="FFFFFF", bold=True)
    thin_black = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for cell in ws[1]:
        cell.fill = fill_head
        cell.font = font_head
        cell.border = thin_black


def _autoajustar_colunas(ws, limite: int = 48):
    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, limite)


def gerar_excel_simples(df_export: pd.DataFrame, sheet_name: str = "dados") -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        ws = writer.sheets[sheet_name[:31]]
        _estilizar_cabecalho(ws)
        _autoajustar_colunas(ws)
    output.seek(0)
    return output.getvalue()


def gerar_excel_confiabilidade(tipo_res: pd.DataFrame, skus_df: pd.DataFrame, detalhe_df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumo = tipo_res.copy()
        if resumo.empty:
            resumo = pd.DataFrame({"mensagem": ["Sem dados para exportar"]})
        resumo.to_excel(writer, sheet_name="Resumo_Tipos", index=False)

        skus = skus_df.copy()
        if skus.empty:
            skus = pd.DataFrame({"mensagem": ["Sem SKUs filtrados"]})
        skus.to_excel(writer, sheet_name="SKUs_Filtrados", index=False)

        detalhe = detalhe_df.copy()
        if detalhe.empty:
            detalhe = pd.DataFrame({"mensagem": ["Sem detalhe por endereço"]})
        detalhe.to_excel(writer, sheet_name="Detalhe_Enderecos", index=False)

        for ws in writer.book.worksheets:
            _estilizar_cabecalho(ws)
            _autoajustar_colunas(ws, limite=60)

        # Coloração por tipo de divergência
        cor_map = {
            "OK": "22C55E",
            "PLAN06_MAIS_RECENTE": "FACC15",
            "OPERACAO_MAIS_RECENTE": "EF4444",
            "DIVERGENTE": "8B5CF6",
            "SEM_MOV_OPERACIONAL": "FB8C00",
            "SEM_PLAN06": "94A3B8",
            "SEM BASE": "CBD5E1",
        }

        for sheet in ["Resumo_Tipos", "SKUs_Filtrados", "Detalhe_Enderecos"]:
            ws = writer.sheets[sheet]
            header_map = {cell.value: cell.column for cell in ws[1]}
            for nome_col in ["tipo_divergencia", "status_validacao"]:
                col_idx = header_map.get(nome_col)
                if not col_idx:
                    continue
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row=row, column=col_idx).value
                    cor = cor_map.get(str(val), None)
                    if cor:
                        ws.cell(row=row, column=col_idx).fill = PatternFill(fill_type="solid", fgColor=cor)
                        ws.cell(row=row, column=col_idx).font = Font(color="FFFFFF" if cor in {"EF4444", "8B5CF6"} else "000000", bold=True)

    output.seek(0)
    return output.getvalue()


def gerar_excel_colorido(df_export: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Heatmap_V15_3", index=False)
        ws = writer.sheets["Heatmap_V15_3"]

        fill_verde = PatternFill(fill_type="solid", fgColor="22C55E")
        fill_amarelo = PatternFill(fill_type="solid", fgColor="FACC15")
        fill_vermelho = PatternFill(fill_type="solid", fgColor="EF4444")
        fill_azul = PatternFill(fill_type="solid", fgColor="1F4E78")
        fill_laranja = PatternFill(fill_type="solid", fgColor="FB8C00")
        font_branca = Font(color="FFFFFF", bold=True)
        font_preta = Font(color="000000", bold=False)
        thin_black = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for cell in ws[1]:
            cell.font = font_branca
            cell.fill = fill_azul
            cell.border = thin_black

        header_map = {cell.value: cell.column for cell in ws[1]}
        col_curva_rep = header_map.get("CURVA_REPOSICAO")
        col_curva_ok = header_map.get("CURVA_CORRETA")
        col_flag = header_map.get("FLAG_URGENTE_LT1D")

        for row in range(2, ws.max_row + 1):
            curva_rep = ws.cell(row=row, column=col_curva_rep).value if col_curva_rep else None
            curva_ok = ws.cell(row=row, column=col_curva_ok).value if col_curva_ok else None
            flag_val = ws.cell(row=row, column=col_flag).value if col_flag else None

            fill = None
            font = font_preta
            if curva_rep == "P":
                fill = fill_verde
            elif curva_rep == "Q":
                fill = fill_amarelo
            elif curva_rep == "R":
                fill = fill_vermelho
                font = font_branca

            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).font = font
                    ws.cell(row=row, column=col).border = thin_black

            if curva_ok == "NAO":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).border = Border(
                        left=Side(style="thick", color="FB8C00"),
                        right=Side(style="thick", color="FB8C00"),
                        top=Side(style="thick", color="FB8C00"),
                        bottom=Side(style="thick", color="FB8C00"),
                    )

            if flag_val in [True, "TRUE", "True", 1]:
                if col_flag:
                    ws.cell(row=row, column=col_flag).fill = fill_laranja
                    ws.cell(row=row, column=col_flag).font = font_branca

        for idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)

    output.seek(0)
    return output.getvalue()


# -----------------------------
# APP
# -----------------------------
st.title("🔥 Heatmap WMS V16 — Visão 2D + 3D")
st.caption("Evolução da base V13 com visão 2D profissional, cockpit 3D e diagnóstico detalhado dos filtros.")

try:
    test_connection()
except Exception as e:
    st.error(f"Erro de conexão com o banco: {e}")
    st.stop()

try:
    empresas = load_empresas()
except Exception as e:
    st.error(f"Erro ao carregar empresas da view vw_mapa_3d_cd: {e}")
    st.stop()

if not empresas:
    st.warning("Nenhuma empresa encontrada na vw_mapa_3d_cd.")
    st.stop()

# Top controls
col_top_1, col_top_2, col_top_3, col_top_4 = st.columns([1, 1, 1, 1])
with col_top_1:
    empresa = st.selectbox("Empresa", options=empresas, index=0)
with col_top_2:
    modo = st.selectbox(
        "Camada do mapa",
        options=["Mapa de Giro", "Pressão Operacional", "Confiabilidade de Dados", "Curva Reposição", "Curva Endereço", "Curva Correta", "Status", "Ocupação"],
        index=0,
    )
with col_top_3:
    vista_3d = st.selectbox(
        "Vista 3D",
        options=["Perspectiva Livre", "Frontal", "Superior", "Lateral", "Isométrica", "Corredor"],
        index=0,
    )
with col_top_4:
    layout_mode = st.selectbox(
        "Modo de layout",
        options=["Compacto", "Explodido por Nível", "Explodido por Prédio", "Explodido Completo"],
        index=0,
    )

try:
    df_base = load_data(empresa)
except Exception as e:
    st.error(f"Erro ao carregar dados: {e}")
    st.stop()

if df_base.empty:
    st.warning("Nenhum dado encontrado para a empresa selecionada.")
    st.stop()

df_base = normalizar_dados(df_base)

try:
    incons_df = load_inconsistencia_saida(int(empresa))
except Exception:
    incons_df = pd.DataFrame(columns=["codigo", "ultima_saida_plan06", "ultima_movimentacao", "dias_plan06", "dias_mov_operacional", "tipo_divergencia", "delta_dias"])

if not incons_df.empty:
    incons_df = incons_df.copy()
    incons_df["codigo"] = pd.to_numeric(incons_df["codigo"], errors="coerce")
    incons_df["ultima_saida_plan06"] = pd.to_datetime(incons_df["ultima_saida_plan06"], errors="coerce")
    incons_df["ultima_movimentacao"] = pd.to_datetime(incons_df["ultima_movimentacao"], errors="coerce")
    incons_df["ULTIMA_SAIDA_PLAN06_TXT"] = incons_df["ultima_saida_plan06"].dt.strftime("%d/%m/%Y %H:%M:%S").fillna("")
    incons_df["ULTIMA_MOV_OPERACIONAL_TXT"] = incons_df["ultima_movimentacao"].dt.strftime("%d/%m/%Y %H:%M:%S").fillna("")
    incons_df = incons_df.rename(columns={
        "codigo": "codproduto",
        "dias_plan06": "DIAS_PLAN06",
        "dias_mov_operacional": "DIAS_MOV_OPERACIONAL",
        "tipo_divergencia": "TIPO_DIVERGENCIA",
        "delta_dias": "DELTA_DIAS_SAIDA",
    })
    df_base = df_base.merge(
        incons_df[["codproduto", "ultima_saida_plan06", "ultima_movimentacao", "ULTIMA_SAIDA_PLAN06_TXT", "ULTIMA_MOV_OPERACIONAL_TXT", "DIAS_PLAN06", "DIAS_MOV_OPERACIONAL", "TIPO_DIVERGENCIA", "DELTA_DIAS_SAIDA"]].drop_duplicates(subset=["codproduto"]),
        on="codproduto",
        how="left",
    )
else:
    df_base["ultima_saida_plan06"] = pd.NaT
    df_base["ultima_movimentacao"] = pd.NaT
    df_base["ULTIMA_SAIDA_PLAN06_TXT"] = ""
    df_base["ULTIMA_MOV_OPERACIONAL_TXT"] = ""
    df_base["DIAS_PLAN06"] = np.nan
    df_base["DIAS_MOV_OPERACIONAL"] = np.nan
    df_base["TIPO_DIVERGENCIA"] = "SEM BASE"
    df_base["DELTA_DIAS_SAIDA"] = np.nan

df_base["TIPO_DIVERGENCIA"] = df_base["TIPO_DIVERGENCIA"].fillna("SEM BASE").apply(normalizar_tipo_divergencia)

df_base["PRESSAO_OPERACIONAL"] = (
    df_base["visitas"].fillna(0) * np.maximum(df_base["gap_curva"].fillna(0), 1)
    + df_base["qtd_saida"].fillna(0) * 0.15
    + np.where(df_base["FLAG_URGENTE_LT1D"], 25, 0)
).round(2)

with st.sidebar:
    st.header("Filtros Operacionais")

    dep_opts = ordenar_natural(df_base["DEP_TXT"].dropna().unique().tolist())
    rua_opts = ordenar_natural(df_base["RUA_TXT"].dropna().unique().tolist())
    predio_opts = ordenar_natural(df_base["PREDIO_TXT"].dropna().unique().tolist())
    nivel_opts = ordenar_natural(df_base["NIVEL_TXT"].dropna().unique().tolist(), reverse=True)
    sala_opts = ordenar_natural(df_base["SALA_TXT"].dropna().unique().tolist())

    dep_sel = st.multiselect("DEP", dep_opts, default=dep_opts)
    rua_sel = st.multiselect("Rua", rua_opts, default=[])
    predio_sel = st.multiselect("Prédio", predio_opts, default=[])
    nivel_sel = st.multiselect("Nível", nivel_opts, default=[])
    sala_sel = st.multiselect("Sala", sala_opts, default=[])

    curva_rep_sel = st.multiselect(
        "Curva Reposição",
        options=sorted([x for x in df_base["CURVA_REPOSICAO"].dropna().unique().tolist() if x != ""]),
        default=[],
    )
    curva_end_sel = st.multiselect(
        "Curva Endereço",
        options=sorted([x for x in df_base["CURVA_ENDERECO"].dropna().unique().tolist() if x != ""]),
        default=[],
    )
    status_sel = st.multiselect(
        "Status do Endereço",
        options=sorted([x for x in df_base["STATUS_ENDERECO"].dropna().unique().tolist() if x != ""]),
        default=[],
    )
    divergencia_sel = st.multiselect(
        "Confiabilidade do dado",
        options=sorted([x for x in df_base["TIPO_DIVERGENCIA"].dropna().unique().tolist() if x != ""]),
        default=[],
    )

    pareto_label = st.selectbox("Pareto", ["Top 20%", "Top 40%", "Top 60%", "Top 80%", "Top 100%"], index=4)
    pareto_map = {"Top 20%": 0.20, "Top 40%": 0.40, "Top 60%": 0.60, "Top 80%": 0.80, "Top 100%": 1.00}
    pareto_pct = pareto_map[pareto_label]

    somente_urgentes = st.checkbox("Somente urgentes < 1 dia", value=False)
    ocultar_invalidos = st.checkbox("Ocultar sem endereço / REC / inválidos", value=True)
    somente_ativos = st.checkbox("Somente endereços ativos", value=True)

    marker_size = st.slider("Tamanho do cubo 3D", min_value=5, max_value=18, value=9)
    opacity = st.slider("Opacidade 3D", min_value=0.30, max_value=1.00, value=0.92)

    st.markdown("---")
    limpar = st.button("Limpar filtros de seleção", use_container_width=True)

if limpar:
    st.rerun()

filtros = {
    "dep": dep_sel,
    "rua": rua_sel,
    "predio": predio_sel,
    "nivel": nivel_sel,
    "sala": sala_sel,
    "curva_rep": curva_rep_sel,
    "curva_end": curva_end_sel,
    "status": status_sel,
    "divergencia": divergencia_sel,
    "somente_urgentes": somente_urgentes,
    "ocultar_invalidos": ocultar_invalidos,
    "somente_ativos": somente_ativos,
    "pareto": pareto_pct,
}

diag_df = build_filter_diagnostics(df_base, filtros)

df = df_base.copy()
if dep_sel:
    df = df[df["DEP_TXT"].isin(dep_sel)]
if rua_sel:
    df = df[df["RUA_TXT"].isin(rua_sel)]
if predio_sel:
    df = df[df["PREDIO_TXT"].isin(predio_sel)]
if nivel_sel:
    df = df[df["NIVEL_TXT"].isin(nivel_sel)]
if sala_sel:
    df = df[df["SALA_TXT"].isin(sala_sel)]
if curva_rep_sel:
    df = df[df["CURVA_REPOSICAO"].isin(curva_rep_sel)]
if curva_end_sel:
    df = df[df["CURVA_ENDERECO"].isin(curva_end_sel)]
if status_sel:
    df = df[df["STATUS_ENDERECO"].isin(status_sel)]
if divergencia_sel:
    df = df[df["TIPO_DIVERGENCIA"].isin(divergencia_sel)]
if somente_urgentes:
    df = df[df["FLAG_URGENTE_LT1D"]]
if ocultar_invalidos:
    df = df[~endereco_invalido_mask(df)]
if somente_ativos:
    df = df[endereco_ativo_mask(df)]

df = aplicar_pareto(df, pareto_pct)

if df.empty:
    st.warning("Os filtros selecionados não retornaram dados.")
    st.subheader("Validador da aplicação dos filtros")
    st.dataframe(diag_df, use_container_width=True, hide_index=True)
    st.stop()

kpis = metricas_resumo(df)
divergentes_kpi = int(df["codproduto"].where(df["TIPO_DIVERGENCIA"].isin(["DIVERGENTE","PLAN06_MAIS_RECENTE","OPERACAO_MAIS_RECENTE","SEM_MOV_OPERACIONAL","SEM_PLAN06"])).dropna().nunique())
col_kpi_1, col_kpi_2, col_kpi_3, col_kpi_4, col_kpi_5, col_kpi_6, col_kpi_7 = st.columns(7)
col_kpi_1.metric("Endereços", f"{kpis['enderecos']:,}".replace(",", "."))
col_kpi_2.metric("Produtos", f"{kpis['produtos']:,}".replace(",", "."))
col_kpi_3.metric("Visitas", f"{int(kpis['visitas']):,}".replace(",", "."))
col_kpi_4.metric("Urgentes", f"{kpis['urgentes']:,}".replace(",", "."))
col_kpi_5.metric("Curva incorreta", f"{kpis['incorretos']:,}".replace(",", "."))
col_kpi_6.metric("Divergências de dado", f"{divergentes_kpi:,}".replace(",", "."))
col_kpi_7.metric("Ocupação média", f"{kpis['ocupacao_media']:.1f}%")

status_diag = diagnostico_status(df)
qtd_sem_status = int(status_diag.loc[status_diag["STATUS_ENDERECO"].eq("SEM STATUS"), "QTD"].sum()) if not status_diag.empty else 0
if qtd_sem_status > 0:
    st.warning(f"{qtd_sem_status} registros estão sem status válido no banco e aparecem como 'SEM STATUS' no mapa.")

aba2d, abamapa, aba3d, abaconf, abaruas, abadiag, abadados = st.tabs([
    "🗺️ Visão 2D",
    "🧱 Mapa de Ocupação",
    "🏗️ Visão 3D",
    "🔎 Confiabilidade de Dados",
    "🚦 Resumo Operacional",
    "🩺 Diagnóstico dos Filtros",
    "📋 Base Filtrada",
])

with aba2d:
    st.subheader("Mapa 2D Profissional")
    metrica_2d = st.radio(
        "Métrica do heatmap 2D",
        options=["visitas", "qtd_saida", "PRESSAO_OPERACIONAL", "perc_ocupacao", "contagem"],
        horizontal=True,
        format_func=lambda x: {
            "visitas": "Visitas",
            "qtd_saida": "Saída",
            "PRESSAO_OPERACIONAL": "Pressão Operacional",
            "perc_ocupacao": "Ocupação Média",
            "contagem": "Itens",
        }[x],
    )
    col2d_1, col2d_2 = st.columns(2)
    with col2d_1:
        st.plotly_chart(gerar_heatmap_rua_nivel(df, metrica_2d), use_container_width=True)
    with col2d_2:
        st.plotly_chart(gerar_mapa_2d_rua_predio(df, "Pressão Operacional" if metrica_2d == "PRESSAO_OPERACIONAL" else "Mapa de Giro"), use_container_width=True)

with abamapa:
    st.subheader("Mapa de Ocupação Clássico")
    st.caption("Visão no padrão painel de ocupação: apenas endereços ativos da base, com ruas, níveis, posições e ocupação geral.")

    if "mapa_fluxo_modo" not in st.session_state:
        st.session_state["mapa_fluxo_modo"] = "Nenhum"

    c_map_1, c_map_2, c_map_3, c_map_4, c_map_5 = st.columns([1.2, 1, 1, 1.05, 1.0])
    with c_map_1:
        localizacao_busca = st.text_input(
            "Localização / endereço para destacar",
            value="",
            placeholder="Ex.: 01-001-02-03-01 ou 001",
            key="mapa_ocupacao_busca",
        )
    with c_map_2:
        somente_status = st.selectbox(
            "Modo do painel",
            options=["Todos", "Somente livres", "Somente vazios", "Somente ocupados", "Somente bloqueados/avariados"],
            index=0,
            key="mapa_ocupacao_modo",
        )
    with c_map_3:
        quebra_painel = st.selectbox(
            "Separação do mapa",
            options=["Geral", "Por DEP", "Por Zona", "DEP + Zona"],
            index=0,
            key="mapa_ocupacao_quebra",
        )
    with c_map_4:
        visao_cor_mapa = st.selectbox(
            "Coloração do mapa",
            options=["Status", "PQR"],
            index=0,
            key="mapa_ocupacao_visao_cor",
        )
    with c_map_5:
        alerta_ocup = st.slider(
            "Alerta de ocupação geral (%)",
            min_value=50,
            max_value=100,
            value=80,
            key="mapa_ocupacao_alerta",
        )

    enderecos_disp = ordenar_natural(df["ENDERECO_LABEL"].dropna().astype(str).unique().tolist())
    c_map_5, c_map_6, c_map_7, c_map_8, c_map_9 = st.columns([1.3, .8, .8, .9, 1.1])
    with c_map_5:
        endereco_sel = st.selectbox("Endereço clicado / selecionado", options=[""] + enderecos_disp, index=0, key="mapa_ocupacao_endereco")
    with c_map_6:
        if st.button("Entrada", use_container_width=True):
            st.session_state["mapa_fluxo_modo"] = "Entrada"
    with c_map_7:
        if st.button("Saída", use_container_width=True):
            st.session_state["mapa_fluxo_modo"] = "Saída"
    with c_map_8:
        if st.button("Limpar realce", use_container_width=True):
            st.session_state["mapa_fluxo_modo"] = "Nenhum"
    with c_map_9:
        pqr_filtro = st.multiselect("Filtro PQR do painel", options=["P", "Q", "R", "SEM PQR"], default=[], key="mapa_ocupacao_pqr")

    # Se o usuário já filtrou a base por Curva Reposição/Endereço, o painel deve
    # automaticamente respeitar isso também na coloração PQR.
    pqr_global = sorted({
        normalizar_pqr(x)
        for x in (list(curva_rep_sel) + list(curva_end_sel))
        if normalizar_pqr(x) in ["P", "Q", "R", "SEM PQR"]
    })
    pqr_ativo = pqr_filtro if pqr_filtro else pqr_global

    modo_fluxo = st.session_state.get("mapa_fluxo_modo", "Nenhum")
    st.caption(
        f"Fluxo ativo no painel: {modo_fluxo}. "
        f"PQR ativo no painel: {', '.join(pqr_ativo) if pqr_ativo else 'nenhum'}. "
        "Quando houver filtro global de Curva ou Filtro PQR do painel, a coloração PQR é aplicada automaticamente."
    )

    mapa_occ_df = preparar_mapa_ocupacao(df)
    total_valid = int(mapa_occ_df["STATUS_MAPA"].isin(STATUSS_ATIVOS_MAPA).sum()) if not mapa_occ_df.empty else 0
    total_ocup = int(mapa_occ_df["STATUS_MAPA"].eq("OCUPADO").sum()) if not mapa_occ_df.empty else 0
    total_livre = int(mapa_occ_df["STATUS_MAPA"].eq("LIVRE").sum()) if not mapa_occ_df.empty else 0
    total_reserv = int(mapa_occ_df["STATUS_MAPA"].eq("RESERVADO").sum()) if not mapa_occ_df.empty else 0
    perc_geral = (total_ocup / total_valid * 100.0) if total_valid else 0.0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Posições ocupadas", f"{total_ocup:,}".replace(",", "."))
    k2.metric("Posições livres", f"{total_livre:,}".replace(",", "."))
    k3.metric("Reservadas", f"{total_reserv:,}".replace(",", "."))
    k4.metric("Ocupação geral", f"{perc_geral:.1f}%")
    k5.metric("Total posições ativas", f"{total_valid:,}".replace(",", "."))

    if perc_geral >= alerta_ocup:
        st.error(f"⚠️ Atenção! O armazém atingiu {perc_geral:.1f}% de ocupação. Risco de congestionamento operacional.")
    else:
        st.success(f"Nível de ocupação sob controle: {perc_geral:.1f}%.")

    visao_cor_entrada = "PQR" if (pqr_ativo or visao_cor_mapa == "PQR") else visao_cor_mapa
    html_mapa, destaque_df = render_mapa_ocupacao_html(
        df,
        somente_status=somente_status,
        localizacao_busca=localizacao_busca,
        endereco_selecionado=endereco_sel,
        modo_fluxo=modo_fluxo,
        quebra_painel=quebra_painel,
        visao_cor=visao_cor_entrada,
        pqr_filtro=pqr_ativo,
    )
    st.components.v1.html(html_mapa, height=980, scrolling=True)

    c_det1, c_det2 = st.columns([1.3, 1])
    with c_det1:
        st.download_button(
            "Baixar Excel do mapa de ocupação",
            data=gerar_planilha_mapa_ocupacao(df),
            file_name=f"mapa_ocupacao_empresa_{empresa}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with c_det2:
        st.caption("Dica: use a busca, o endereço selecionado, os botões Entrada/Saída e a coloração por Status ou PQR. Se houver filtro global em Curva Reposição/Endereço, o painel muda automaticamente para PQR.")

    if localizacao_busca or endereco_sel:
        st.markdown("### Detalhe da localização pesquisada")
        if destaque_df.empty:
            st.info("Nenhuma posição do mapa clássico corresponde ao texto informado.")
        else:
            st.dataframe(destaque_df, use_container_width=True, hide_index=True)


with aba3d:
    st.subheader("Cockpit 3D")
    vista_norm = vista_3d.replace("Perspectiva Livre", "Perspectiva Livre")
    st.plotly_chart(
        montar_figura_3d(df, modo, vista_norm, layout_mode, marker_size, opacity),
        use_container_width=True,
    )


with abaconf:
    st.subheader("Confiabilidade de Dados — V17.3")
    st.caption("Camada de auditoria entre PLAN06 e movimentação operacional. A V17.3 adiciona exportação inteligente com múltiplas abas para tratamento no C5.")

    incons_base = (
        df[[
            "codproduto", "descproduto", "TIPO_DIVERGENCIA", "DELTA_DIAS_SAIDA",
            "ULTIMA_SAIDA_PLAN06_TXT", "ULTIMA_MOV_OPERACIONAL_TXT",
            "DIAS_PLAN06", "DIAS_MOV_OPERACIONAL", "STATUS_ENDERECO",
            "ENDERECO_LABEL", "RUA_TXT", "PREDIO_TXT", "NIVEL_TXT", "SALA_TXT",
            "qtd_saida", "visitas", "PRESSAO_OPERACIONAL"
        ]]
        .drop_duplicates(subset=["codproduto", "ENDERECO_LABEL"])
        .copy()
    )
    incons_base = incons_base.rename(columns={
        "codproduto": "codigo",
        "descproduto": "descricao",
        "TIPO_DIVERGENCIA": "tipo_divergencia",
        "DELTA_DIAS_SAIDA": "delta_dias",
        "ULTIMA_SAIDA_PLAN06_TXT": "ultima_saida_plan06",
        "ULTIMA_MOV_OPERACIONAL_TXT": "ultima_movimentacao_operacional",
        "STATUS_ENDERECO": "status_endereco",
        "ENDERECO_LABEL": "endereco",
        "RUA_TXT": "rua",
        "PREDIO_TXT": "predio",
        "NIVEL_TXT": "nivel",
        "SALA_TXT": "sala",
        "qtd_saida": "qtd_saida_mapa",
        "visitas": "visitas_mapa",
        "PRESSAO_OPERACIONAL": "score_pressao_mapa",
    })
    incons_sku = (
        incons_base.groupby(["codigo", "descricao", "tipo_divergencia", "delta_dias", "ultima_saida_plan06", "ultima_movimentacao_operacional", "DIAS_PLAN06", "DIAS_MOV_OPERACIONAL"], as_index=False)
        .agg(
            qtd_enderecos=("endereco", "nunique"),
            score_pressao=("score_pressao_mapa", "sum"),
            visitas=("visitas_mapa", "sum"),
            qtd_saida=("qtd_saida_mapa", "sum"),
        )
    )
    tipo_opts = ["OK", "PLAN06_MAIS_RECENTE", "OPERACAO_MAIS_RECENTE", "DIVERGENTE", "SEM_MOV_OPERACIONAL", "SEM_PLAN06", "SEM BASE"]
    tipo_res = (
        incons_sku.groupby("tipo_divergencia", as_index=False)
        .agg(qtd_skus=("codigo", "nunique"))
        .sort_values("qtd_skus", ascending=False)
    )
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("SKUs divergentes", f"{int(incons_sku[incons_sku['tipo_divergencia'].ne('OK')]['codigo'].nunique()):,}".replace(",", "."))
    c2.metric("SKUs OK", f"{int(incons_sku[incons_sku['tipo_divergencia'].eq('OK')]['codigo'].nunique()):,}".replace(",", "."))
    c3.metric("Maior delta (dias)", f"{int(pd.to_numeric(incons_sku['delta_dias'], errors='coerce').fillna(0).max() if not incons_sku.empty else 0):,}".replace(",", "."))
    c4.metric("SKUs sem base operacional", f"{int(incons_sku[incons_sku['tipo_divergencia'].eq('SEM_MOV_OPERACIONAL')]['codigo'].nunique()):,}".replace(",", "."))

    g1, g2 = st.columns([1, 1.2])
    with g1:
        st.markdown("### Resumo por tipo de divergência")
        st.dataframe(tipo_res, use_container_width=True, hide_index=True)
    with g2:
        fig_tipo = px.bar(
            tipo_res,
            x="tipo_divergencia",
            y="qtd_skus",
            color="tipo_divergencia",
            color_discrete_map={k: cor_confiabilidade(k) for k in tipo_opts},
            text="qtd_skus",
            labels={"tipo_divergencia": "Tipo", "qtd_skus": "SKUs"},
        )
        fig_tipo.update_layout(height=360, margin=dict(l=0, r=0, t=10, b=0), xaxis_title="", yaxis_title="SKUs", showlegend=False)
        st.plotly_chart(fig_tipo, use_container_width=True)

    cf1, cf2, cf3 = st.columns([1.2, 1, 0.8])
    with cf1:
        tipos_sel_conf = st.multiselect("Tipos para analisar", options=tipo_opts, default=[x for x in tipo_opts if x != "OK"], key="tipos_conf_sel")
    with cf2:
        delta_min = st.number_input("Delta mínimo de dias", min_value=0, max_value=9999, value=5, step=1, key="delta_min_conf")
    with cf3:
        preview_conf = st.selectbox("Prévia", options=[50, 100, 200, 500], index=1, key="preview_conf")

    conf_view = incons_sku.copy()
    if tipos_sel_conf:
        conf_view = conf_view[conf_view["tipo_divergencia"].isin(tipos_sel_conf)]
    conf_view = conf_view[pd.to_numeric(conf_view["delta_dias"], errors="coerce").fillna(0) >= float(delta_min)]
    conf_view = conf_view.sort_values(["delta_dias", "score_pressao", "visitas"], ascending=[False, False, False])

    st.markdown("### SKUs para exportar e tratar no C5")
    st.dataframe(conf_view.head(int(preview_conf)), use_container_width=True, hide_index=True)
    detalhe_end = incons_base.copy()
    if tipos_sel_conf:
        detalhe_end = detalhe_end[detalhe_end["tipo_divergencia"].isin(tipos_sel_conf)]
    detalhe_end = detalhe_end[pd.to_numeric(detalhe_end["delta_dias"], errors="coerce").fillna(0) >= float(delta_min)]

    cexp1, cexp2 = st.columns(2)
    with cexp1:
        st.download_button(
            "📥 Exportar divergências premium (Excel)",
            data=gerar_excel_confiabilidade(tipo_res, conf_view, detalhe_end),
            file_name=f"divergencias_v17_3_{empresa}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with cexp2:
        st.download_button(
            "📥 Exportar SKUs filtrados (Excel simples)",
            data=gerar_excel_simples(conf_view, sheet_name="divergencias"),
            file_name=f"divergencias_skus_{empresa}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with st.expander("Detalhe por endereço / mapa"):
        st.dataframe(detalhe_end.head(1000), use_container_width=True, hide_index=True)

with abaruas:
    st.subheader("Resumo executivo por rua")
    resumo_rua_df = resumo_por_rua(df)
    top_criticos_df = top_enderecos_criticos(df, top_n=25)
    c1, c2 = st.columns([1.15, 1])
    with c1:
        st.dataframe(resumo_rua_df, use_container_width=True, hide_index=True)
    with c2:
        st.dataframe(top_criticos_df, use_container_width=True, hide_index=True)

with abadiag:
    st.subheader("Validador da aplicação dos filtros")
    st.dataframe(diag_df, use_container_width=True, hide_index=True)
    etapa_critica = diag_df.loc[diag_df["qtd_linhas"].eq(0), "etapa"]
    if not etapa_critica.empty:
        st.error(f"A base zera na etapa: {etapa_critica.iloc[0]}")
    else:
        st.success("Os filtros estão consistentes. A base final permanece com dados.")

    st.markdown("### Filtro SQL de validação rápida")
    deps_sql = ", ".join([f"'{x}'" for x in dep_sel]) if dep_sel else ""
    ruas_sql = ", ".join([f"'{x}'" for x in rua_sel]) if rua_sel else ""
    sql_validacao = f"""
SELECT
    nroempresa,
    dep,
    rua,
    predio,
    nivel,
    sala,
    codproduto,
    descproduto,
    visitas,
    qtd_saida,
    curva_reposicao,
    curva_endereco,
    status_endereco
FROM vw_mapa_3d_cd
WHERE nroempresa = {int(empresa)}
{f"  AND LPAD(CAST(dep AS CHAR), 2, '0') IN ({deps_sql})" if deps_sql else ""}
{f"  AND LPAD(CAST(rua AS CHAR), 3, '0') IN ({ruas_sql})" if ruas_sql else ""}
LIMIT 200;
""".strip()
    st.code(sql_validacao, language="sql")

with abadados:
    st.subheader("Base filtrada")
    export_cols = [
        "nroempresa", "DEP_TXT", "RUA_TXT", "PREDIO_TXT", "NIVEL_TXT", "SALA_TXT",
        "ENDERECO_LABEL", "codproduto", "descproduto", "visitas", "qtd_saida", "media_dia_cx",
        "dias_sem_reposicao", "CURVA_REPOSICAO", "CURVA_ENDERECO", "CURVA_CORRETA",
        "perc_ocupacao", "STATUS_ENDERECO", "FLAG_URGENTE_LT1D", "PRESSAO_OPERACIONAL",
        "TIPO_DIVERGENCIA", "DELTA_DIAS_SAIDA", "ULTIMA_SAIDA_PLAN06_TXT", "ULTIMA_MOV_OPERACIONAL_TXT",
        "justificativa", "acao_recomendada"
    ]
    export_df = df[export_cols].copy().rename(columns={
        "DEP_TXT": "DEP",
        "RUA_TXT": "RUA",
        "PREDIO_TXT": "PREDIO",
        "NIVEL_TXT": "NIVEL",
        "SALA_TXT": "SALA",
    })
    st.dataframe(export_df, use_container_width=True, hide_index=True)
    st.download_button(
        "Baixar Excel colorido",
        data=gerar_excel_colorido(export_df),
        file_name=f"heatmap_wms_v15_2_empresa_{empresa}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.markdown("---")
st.markdown(
    """
**Leitura operacional V15.5**
- **Visão 2D**: leitura tática por rua, prédio e nível.
- **Visão 3D**: cockpit espacial do layout com camadas de giro, ocupação, pressão e criticidade.
- **Validador de filtros**: mostra exatamente em qual etapa a base reduz até zerar.
- **Pressão Operacional**: score composto por visitas, gap de curva, saída e urgência.
"""
)

@st.cache_data(ttl=300, show_spinner=False)
def load_inconsistencia_saida(nroempresa: int):
    sql = text(
        """
        SELECT
            nroempresa,
            codigo,
            descricao,
            ultima_saida_plan06,
            ultima_movimentacao,
            dias_plan06,
            dias_mov_operacional,
            tipo_divergencia,
            delta_dias
        FROM vw_inconsistencia_saida
        WHERE nroempresa = :nroempresa
        """
    )
    with get_engine().connect() as conn:
        return pd.read_sql(sql, conn, params={"nroempresa": str(nroempresa)})



